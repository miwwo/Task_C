import vk_api
from vk_api.keyboard import VkKeyboard, VkKeyboardColor
from vk_api.longpoll import VkLongPoll, VkEventType
from vk_api.utils import get_random_id
import datetime
import requests
from bs4 import BeautifulSoup
import pickle
# import dateutil
import json
import os
import openpyxl
import random
import matplotlib.pyplot as plt
import matplotlib.ticker as mticker
import matplotlib
import PIL.Image as Image
import re
from math import ceil, floor
from range_key_dict import RangeKeyDict

class UserInfo:
    def __init__(self):
        self.name = ""
        self.group = ""
        self.course = ""
        self.answering = ""
        self.other_group = ""
        self.buf = []

#для погоды
# время суток
day_time = RangeKeyDict({(5, 11): 0, (11, 16): 1, (16, 23): 2, (23, 24): 3, (0, 5): 3})
day_name =  {0: "утро", 1: "день", 2: "вечер", 3: "ночь"}
# характеристика ветра
wind = RangeKeyDict({(0, 0.3): 'штиль', (0.3, 1.6): 'тихий',
        (1.6, 3.4): 'легкий', (3.4, 5.5): 'слабый', (5.5, 8): 'умеренный',
        (8, 10.8): 'свежий', (10.8, 13.9): 'сильный', (13.9, 17.2): 'крепкий',
        (17.2, 20.8): 'очень крепкий', (20.8, 24.5): 'шторм', (24.5, 28.5): 'сильный шотрм',
        (28.5, 32.7): 'жестокий шторм', (32.7, 200): 'ураган'})
wind_direction = RangeKeyDict({(0, 22.5):'северный',
(22.5, 67.5):'северо-восточный',
(67.5, 112.5):'восточный',
(112.5, 157.5):'юго-восточный',
(157.5, 202.5):'южный',
(202.5, 247.5):'юго-западный',
(247.5, 292.5):'западный',
(292.5, 337.5):'северо-западный',
(337.5, 360.1):'северный'})
#для коронавируса
reg_dict = {}
reg_list = []
page_reg = requests.get('https://coronavirusstat.ru/country/russia/') # адрес старницы со статистикой
region_soup = BeautifulSoup(page_reg.text, "html.parser")
a = (region_soup.find_all('div', class_="p-1 col-5"))
for i in a:
    reg_list.append(' '.join(i.text.split()[1:]))
    reg_dict[' '.join(i.text.split()[1:])] = 'https://coronavirusstat.ru' + i.find('a').get('href')

def create_keyboard(response):
    if response == "name":
        keyboard = VkKeyboard(one_time=True)
        keyboard.add_button('Да', color=VkKeyboardColor.POSITIVE)
        keyboard.add_line()
        keyboard.add_button('Нет', color=VkKeyboardColor.NEGATIVE)
        keyboard = keyboard.get_keyboard()
        vk.messages.send(
            user_id=id,
            random_id=get_random_id(), message="Выберите пункт или отправьте сообщение",
            keyboard=keyboard)
    if response == "start":
        keyboard = VkKeyboard(one_time=True)
        keyboard.add_button('Расписание', color=VkKeyboardColor.PRIMARY)
        keyboard.add_line()
        keyboard.add_button('Погода', color=VkKeyboardColor.PRIMARY)
        keyboard.add_line()
        keyboard.add_button('Корона', color=VkKeyboardColor.PRIMARY)
        keyboard = keyboard.get_keyboard()
        vk.messages.send(
            user_id=id,
            random_id=get_random_id(),
            message="Главное меню", keyboard=keyboard)
    if response == "raspis":
        keyboard = VkKeyboard(one_time=True)
        keyboard.add_button('сегодня', color=VkKeyboardColor.POSITIVE)
        keyboard.add_button('завтра', color=VkKeyboardColor.NEGATIVE)
        keyboard.add_line()  # переход на вторую строку
        keyboard.add_button('эту неделю', color=VkKeyboardColor.PRIMARY)
        keyboard.add_button('следующую неделю', color=VkKeyboardColor.PRIMARY)
        keyboard.add_line()
        keyboard.add_button('Какая неделя?', color=VkKeyboardColor.SECONDARY)
        keyboard.add_button('Какая группа?', color=VkKeyboardColor.SECONDARY)
        # keyboard.add_line()
        # keyboard.add_button('Закрыть клавиатуру', color=VkKeyboardColor.SECONDARY)
        keyboard = keyboard.get_keyboard()
        vk.messages.send(
            user_id=id,
            random_id=get_random_id(),
            message="Расписание на...", keyboard=keyboard)
    if response == "teachers":
        keyboard = VkKeyboard(one_time=True)
        keyboard.add_button(users[id].buf[0], color=VkKeyboardColor.PRIMARY)
        for i in range(1,len(users[id].buf)):
            keyboard.add_line()
            keyboard.add_button(users[id].buf[i], color=VkKeyboardColor.PRIMARY)
        keyboard = keyboard.get_keyboard()
        vk.messages.send(
            user_id=id,
            random_id=get_random_id(),
            message="Найдено несколько преподавателей, какого вы имели в виду?", keyboard=keyboard)
    if response == "teach_raspis":
        keyboard = VkKeyboard(one_time=True)
        keyboard.add_button('сегодня', color=VkKeyboardColor.POSITIVE)
        keyboard.add_button('завтра', color=VkKeyboardColor.NEGATIVE)
        keyboard.add_line()  # переход на вторую строку
        keyboard.add_button('эту неделю', color=VkKeyboardColor.PRIMARY)
        keyboard.add_button('следующую неделю', color=VkKeyboardColor.PRIMARY)
        keyboard = keyboard.get_keyboard()
        vk.messages.send(
            user_id=id,
            random_id=get_random_id(),
            message="Расписание на...", keyboard=keyboard)
    if response == "weather":
        keyboard = VkKeyboard(one_time=True)
        keyboard.add_button("сейчас", color=VkKeyboardColor.PRIMARY)
        keyboard.add_button("сегодня", color=VkKeyboardColor.POSITIVE)
        keyboard.add_button("завтра", color=VkKeyboardColor.POSITIVE)
        keyboard.add_line()
        keyboard.add_button("на 5 дней", color=VkKeyboardColor.POSITIVE)
        keyboard = keyboard.get_keyboard()
        vk.messages.send(
            user_id=id,
            random_id=get_random_id(),
            message="Выберите интересующий период ...", keyboard=keyboard)
def send_message(id, msg):
    try:
        vk.messages.send(
            user_id=id,
            random_id=get_random_id(),
            message=msg)  # chat_id = event.chat_id
    except BaseException:
        return


def parsing():
    for i in range (1,4):
        if not os.path.isfile(f"raspisanie{i}.xlsx"):
            page = requests.get("https://www.mirea.ru/schedule/")
            soup = BeautifulSoup(page.text, "html.parser")
            result = soup.find("div", {'class': "rasspisanie"}). \
                find(string="Институт информационных технологий"). \
                find_parent("div"). \
                find_parent("div"). \
                find_all("a", class_="uk-link-toggle")[i - 1].get('href')
            # find_all("a",class_="uk-link-toggle")     #"uk-card uk-card-default uk-card-body uk-card-small uk-text-center uk-box-shadow-hover-small")
            with open(f"raspisanie{i}.xlsx", "wb") as table:
                resp = requests.get(result)
                table.write(resp.content)



def is_group_name(msg): # ИКБО-08-21
    msg = msg.strip()
    #book = openpyxl.load_workbook(f"raspisanie{users[id].course}.xlsx")  # открытие файла
    try:
        book = openpyxl.load_workbook(f"raspisanie{int(str(datetime.datetime.today())[:4]) - int('20' + msg[8:])}.xlsx")  # открытие файла
        sheet = book.active  # активный лист
        num_cols = sheet.max_column  # количество столбцов
        # num_rows = sheet.max_row  # количество строк
        for i in range(1, num_cols + 1):
            group = sheet.cell(row=2, column=i).value
            group = str(group).strip()
            if group == msg.upper():
                return True
        return False
    except BaseException and ValueError:
        send_message(id,"Ошибка, проверьте ввод группы")


def raspisanie(group):  # получить полное расписание для группы
    # create_keyboard("raspis")
    group = group.upper()
    group = group.strip()
    if os.path.isfile(f'{group}.json'):
        with open(f'{group}.json', 'r') as j:
            group_days = json.load(j)
            return group_days
    else:
        group_days = {}
        #print('20' + group[8:])
        try:
            book = openpyxl.load_workbook(f"raspisanie{int(str(datetime.datetime.today())[:4]) - int('20' + group[8:])}.xlsx")  # открытие файла
        except BaseException and ValueError:
            send_message(id, "Ошибка, файл группы не найден")
            return
        book = openpyxl.load_workbook(
            f"raspisanie{int(str(datetime.datetime.today())[:4]) - int('20' + group[8:])}.xlsx")  # открытие файла
        sheet = book.active  # активный лист
        num_cols = sheet.max_column  # количество столбцов
        num_rows = sheet.max_row
        for g in range(1, num_cols + 1):
            f_group = sheet.cell(row=2, column=g).value
            f_group = str(f_group).strip()
            daycolumn = 0
            if f_group == group:
                strin = -8
                for o in range(1, num_cols + 1):
                    day = sheet.cell(row=2, column=g + o).value
                    if day == "День недели":
                        daycolumn = g+o
                        break
                for i in range(6):
                    # cell = sheet.cell(row=4, column=g-5).value
                    strin += 12
                    day = sheet.cell(row=strin, column=daycolumn).value
                    lesson = []
                    for j in range(12):
                        lesson.append(sheet.cell(row=j + strin, column=g).value)
                        if not lesson[j]:
                            lesson[j] = "—"
                        el = ""
                        for k in range(1, 4):
                            if sheet.cell(row=j + strin, column=g + k).value != "":
                                el = sheet.cell(row=j + strin, column=g + k).value
                            if not el:
                                break
                            else:
                                lesson[j] += ", " + str(el)
                    group_days[day] = lesson
                break
        #print(group_days)
        with open(f'{group}.json', 'w') as j:
            json.dump(group_days, j)
        return group_days


def week_number(d=datetime.date.today()):
    s = int(d.strftime("%V")) - int(datetime.date(2022, 2, 9).strftime("%V")) + 1
    return s


def week_day(d=datetime.datetime.today().weekday()):
    if d == 0:
        return "ПОНЕДЕЛЬНИК"
    elif d == 1:
        return "ВТОРНИК"
    elif d == 2:
        return "СРЕДА"
    elif d == 3:
        return "ЧЕТВЕРГ"
    elif d == 4:
        return "ПЯТНИЦА"
    elif d == 5:
        return "СУББОТА"
    elif d == 6:
        return "ВОСКРЕСЕНЬЕ"
    else:
        return "НЕТ"


def is_sunday(d=datetime.datetime.today().weekday()):
    if week_day(d) == "ВОСКРЕСЕНЬЕ":
        send_message(id, "В воскресенье пар нет")
        users[id].answering = ""
        return True
    else:
        return False


def print_raspis(d=datetime.datetime.today(),now = True,teacher = ""):
    if not teacher:
        rasp_days = raspisanie(users[id].other_group)
    else:
        rasp_days = teacher_rasp(teacher)
    if week_number(d) % 2 != 0:
        if now:
            send_rasp = week_day(d.weekday()) + " " + d.strftime("%d.%m") + "\n"
        else:
            send_rasp = "Нечётная неделя: \n"
        count = 0
        for i in range(0, 12, 2):
            count += 1
            send_rasp += str(count) + ")" + rasp_days[week_day(d.weekday())][i] + "\n"
    else:
        if now:
            send_rasp = week_day(d.weekday()) + " " + d.strftime("%d.%m") + "\n"
        else:
            send_rasp = "Чётная неделя: \n"
        count = 0
        for i in range(1, 12, 2):
            count += 1
            send_rasp += str(count) + ")" + rasp_days[week_day(d.weekday())][i] + "\n"
    return send_rasp


def find_teacher(teacher):
    flag = False
    teacher += " "
    full_teacher = []
    for i in range(1,4):
        try:
            book = openpyxl.load_workbook(f"raspisanie{i}.xlsx")
        except BaseException:
            send_message(id,"Не удалось открыть таблицу")
            return
        sheet = book.active  # активный лист
        num_cols = sheet.max_column  # количество столбцов
        for j in range(1,num_cols+1):
            cell = sheet.cell(row=3, column=j).value
            if cell == "ФИО преподавателя":
                for o in range(4,76):
                    cell = sheet.cell(row=o, column=j).value
                    if cell:
                        #print(teacher, cell, sep = " ")
                        if teacher in str(cell):
                            if full_teacher != [] and cell[cell.find(teacher):cell.find(teacher)+len(teacher) + 4] not in full_teacher:
                                flag = True
                            if cell[cell.find(teacher):cell.find(teacher)+len(teacher) + 4] not in full_teacher:
                                full_teacher.append(cell[cell.find(teacher):cell.find(teacher)+len(teacher) + 4])
    print(full_teacher)
    return full_teacher


def teacher_rasp(teacher):
    rasp = {}
    if os.path.isfile(f'{teacher}.json'):
        with open(f'{teacher}.json', 'r') as j:
            rasp = json.load(j)
            return rasp
    for i in range(1, 4):
        #print(1)
        try:
            book = openpyxl.load_workbook(f"raspisanie{i}.xlsx")
        except BaseException:
            send_message(id, "Не удалось открыть таблицу")
            return
        book = openpyxl.load_workbook(f"raspisanie{i}.xlsx")
        sheet = book.active  # активный лист
        num_cols = sheet.max_column  # количество столбцов
        num_rows = sheet.max_row
        for g in range(1, num_cols + 1):
            cell = sheet.cell(row=3, column=g).value
            daycolumn = 0
            lesson = []
            if cell == "ФИО преподавателя":
                group = sheet.cell(row=2, column=g - 2).value
                strin = -8
                for o in range(1, num_cols + 1):
                    day = sheet.cell(row=2, column=g - o).value
                    if day == "День недели":
                        daycolumn = g - o
                        break
                for i in range(6):
                    # cell = sheet.cell(row=4, column=g-5).value
                    strin += 12
                    #print(daycolumn)
                    day = sheet.cell(row=strin, column=daycolumn).value
                    lesson = []
                    for j in range(12):
                        cell = sheet.cell(row=strin+j, column=g).value
                        #print(teacher, cell)
                        if teacher in str(cell):
                            lesson.append(sheet.cell(row=j + strin, column=g-2).value)
                            el1 = sheet.cell(row=j + strin, column=g - 1).value
                            el2 = sheet.cell(row=j + strin, column=g + 1).value
                            if not el1:
                                el1 = " "
                            if not el2:
                                el2 = " "
                            el = group + ", " + el1 + ", " + el2
                            lesson[j] += ", " + str(el)
                        else:
                            lesson.append("—")
                        #if not lesson[j]:
                         #   lesson[j] = "—"
                    if day in rasp:
                        for l in range(12):
                            if rasp[day][l] == "—":
                                rasp[day][l] = lesson[l]
                    else:
                        rasp[day] = lesson
    with open(f'{teacher}.json', 'w') as j:
        json.dump(rasp, j)
    return rasp


def recieve_answer(ans_type):
    if ans_type == "start_choosing":
        if message == "нет":
            users[id].name = " "
            create_keyboard("start")
            users[id].answering = ""
        else:
            send_message(id, "Как соизволите к вам обращаться?")
            users[id].answering = "name"
    if ans_type == "name":
        users[id].name = event.text
        users[id].answering = ""
        send_message(id, "Ваше имя внесено в базу данных, " + users[id].name)
        with open('data.pickle', 'wb') as f:
            pickle.dump(users, f)
        return
    # if ans_type == "group":
    #   users[id].group = message
    if ans_type == "group":
        if (int(message[0]) >= 1 and int(message[0]) <= 3):
            users[id].course = message[0]
            parsing()
        else:
            send_message(id, "Ошибка ввода, курс от 1 до 3, цифрами")
            send_message(id,
                         "Введите свой курс обучения и группу через пробел (курс от 1 до 3 цифрами). Например:\n1 ИКБО-08-21")
            return
        send_message(id, "Производится поиск...")
        if is_group_name(message[2:]):
            users[id].group = message[2:].upper()
            users[id].other_group = users[id].group
            send_message(id,
                         "Курс вашего обучения и ваша группа внесены в базу данных. Чтобы получить расписание, снова введите слово 'расписание'")
            users[id].answering = ""
            with open('data.pickle', 'wb') as f:
                pickle.dump(users, f)
        else:
            send_message(id, "Группа не найдена, проверьте ввод")
            send_message(id,
                         "Введите свой курс обучения и группу через пробел (курс от 1 до 3 цифрами). Например:\n1 ИКБО-08-21")
        return
    if "wait_raspis" in ans_type:
        users[id].answering = ""
        print(ans_type[11:])
        if message == "сегодня":
            if is_sunday():
                return
            d = datetime.date.today()
            send_rasp = print_raspis(d,True,ans_type[11:])
            send_message(id, send_rasp)
            users[id].other_group = users[id].group
        if message == "завтра":
            d = datetime.date.today() + datetime.timedelta(days=1)
            if is_sunday(d.weekday()):
                return
            send_rasp = print_raspis(d,True,ans_type[11:])
            send_message(id, send_rasp)
            users[id].other_group = users[id].group
        if message == "эту неделю":
            d = datetime.date.today()
            for j in range(7):
                if week_day(d.weekday()) == "ПОНЕДЕЛЬНИК":
                    for i in range(6):
                        send_rasp = print_raspis(d,True,ans_type[11:])
                        send_message(id, send_rasp)
                        d += datetime.timedelta(days=1)
                    break
                else:
                    d -= datetime.timedelta(days=1)
            users[id].other_group = users[id].group
        if message == "следующую неделю":
            d = datetime.date.today() + datetime.timedelta(days=1)
            for j in range(7):
                if week_day(d.weekday()) == "ПОНЕДЕЛЬНИК":
                    for i in range(6):
                        send_rasp = print_raspis(d,True,ans_type[11:])
                        send_message(id, send_rasp)
                        d += datetime.timedelta(days=1)
                    break
                else:
                    d += datetime.timedelta(days=1)
            users[id].other_group = users[id].group
        if "бот" in message:
            d = week_day(0)
            if message[4:] == "воскресенье":
                is_sunday(6)
                return
            flag = False
            for i in range (6):
                if d.lower() == message[4:]:
                    flag = True
                    w = datetime.datetime.today()
                    for j in range (7):
                        if week_day(w.weekday()) == d:
                            send_rasp = print_raspis(w,False)
                            send_message(id, send_rasp)
                            break
                        w += datetime.timedelta(days=1)
                    w += datetime.timedelta(days=7)
                    send_rasp = print_raspis(w,False)
                    send_message(id, send_rasp)
                    break
                d = week_day(i+1)

            if not flag:
                if message[8] == "-" and message[11] == "-": #икбо-08-21
                    if is_group_name(message[3:]):
                        # print(int(str(datetime.datetime.today())[:4]) - int("20" + message[12:]))
                        users[id].answering = "wait_raspis"
                        send_message(id,"Расписание для группы " + message[3:].upper())
                        create_keyboard("raspis")
                        users[id].other_group = message[3:]
                        raspisanie(message[3:])

                    else:
                        send_message(id,"Группа не найдена")
                        return
    if ans_type == "teacher":
        users[id].answering = ""
        teacher = event.text[6:].capitalize()
        teachers = find_teacher(teacher)
        if teachers == []:
            send_message(id,"Преподаватель не найден")
            return
        if len(teachers) > 1:
            users[id].answering = "some_teachers"
            users[id].buf = teachers
            create_keyboard("teachers")
        else:
            #send_raspis = teacher_rasp(teachers[0])
            users[id].answering = 'wait_raspis' + teachers[0]
            create_keyboard("teach_raspis")
        return
    if ans_type == 'some_teachers':
        users[id].answering = ""
        #send_raspis = teacher_rasp(event.text)
        users[id].answering = 'wait_raspis' + event.text
        create_keyboard("teach_raspis")
        return
    if ans_type == "weather":
        if message == "сейчас":
            weather_now()
        elif message == "сегодня":
            send_message(id,"Получение информации...")
            weather_today()
        elif message == "завтра":
            send_message(id, "Получение информации...")
            weather_tomorrow()
        elif message == "на 5 дней":
            send_message(id, "Получение информации...")
            weather5()
        users[id].answering = ""

#скопировал коронавирус у Насти, надо попытаться адаптировать или переписать
def send_photo(user_id, img_req, message = None):
    upload = vk_api.VkUpload(vk_session)
    photo = upload.photo_messages(img_req)[0]
    owner_id = photo['owner_id']
    photo_id = photo['id']
    attachment = f'photo{owner_id}_{photo_id}'
    post = {'user_id': user_id, 'random_id': 0, "attachment": attachment}
    if message != None:
        post['message'] = message
    vk_session.method('messages.send', post)
#статистика по ковиду в россии
def corona_rus():
    page = requests.get('https://coronavirusstat.ru/country/russia/') # адрес страницы со статистикой
    soup = BeautifulSoup(page.text, "html.parser")
    # текстовое сообщение со статистикой
    information = "По состоянию на " + soup.find('strong').text + '\n'
    numbers = soup.find_all('div', class_="col col-6 col-md-3 pt-4")
    for i in numbers:
        number = i.find('b')
        information += number.find_next().text.lower() + ": " + number.text + " (" + i.contents[1].text.replace('(', 'за ') + '\n'
    corona_graph()
    send_photo(event.user_id, 'covid.png', information)
# график
def corona_graph():
    page = requests.get('https://coronavirusstat.ru/country/russia/') # адрес страницы со статистикой
    soup1 = BeautifulSoup(page.text, "html.parser")
    stat_info = soup1.find('table').find('tbody')
    dates = []
    arr1, arr2, arr3 = [], [], []
    for i in range(10):
        dates.append(stat_info.find_all('th')[i].text)
        arr1.append(int(stat_info.find_all('tr')[i].find_all('td')[0].text.split()[0])) # активных
        arr2.append(arr1[i] + int(stat_info.find_all('tr')[i].find_all('td')[1].text.split()[0])) # вылечено
        arr3.append(arr2[i] + int(stat_info.find_all('tr')[i].find_all('td')[2].text.split()[0])) # умерло
    dates.reverse()
    arr1.reverse()
    arr2.reverse()
    arr3.reverse()
    matplotlib.use("TkAgg")
    fig, ax = plt.subplots()
    plt.plot(dates, arr1, "black")
    plt.plot(dates, arr2, "red")
    plt.ylim([0, 20000000])
    plt.title("Россия - детальная статистика - коронавирус")
    plt.plot(dates, arr3, label="умерло", color = 'red')
    plt.fill_between(dates, arr3, color = 'red')
    plt.plot(dates, arr2, label="вылечено", color = 'green')
    plt.fill_between(dates, arr2, color = 'green')
    plt.plot(dates, arr1, label="активных", color = 'yellow')
    plt.fill_between(dates, arr1, color = 'yellow')
    plt.xticks(rotation = 25, fontsize = 7)
    plt.yticks(fontsize = 7)
    plt.gca().yaxis.set_major_formatter(mticker.FormatStrFormatter('%d'))
    plt.legend()
    plt.grid(True)
    fig.savefig('covid.png')
# статистика по региону
def corona_reg(rname):
    reg_name = check_name(rname)
    if reg_name == -1:
        send_message(event.user_id, f"Регион с именем '{rname}' не был найден. Попробуйте ввести название региона иным способом.")
    else:
        page = requests.get(reg_dict[reg_name]) # адрес старницы со статистикой
        soup_reg = BeautifulSoup(page.text, "html.parser")
        information = "По состоянию на " + soup_reg.find('strong').text+ '\nРегион: ' + reg_name + '\n'
        numbers = soup_reg.find_all('div', class_="col col-6 col-md-3 pt-4")
        for i in numbers:
            number = i.find('b')
            information += number.find_next().text.lower() + ": " + number.text + " (" + i.contents[1].text.replace('(', 'за ') + '\n'
        send_message(event.user_id, information)
 #name - название региона вводимое пользователем, ф-ия изменяет название таким образом, чтобы оно соответствовало названию из словаря с регионами
def check_name(name):
    regions = []
    d = [ 'автономная обл.', 'обл.','край','республика','автономный округ', ')']
    for i in range(len(reg_list)):
        regions.append(reg_list[i].lower())
        for j in d:
            regions[i] = regions[i].replace(j, '')
        regions[i] = regions[i].replace(' — ', ' (').replace('  - ', ' (').lstrip().rstrip().split(' (')
    for i in range(len(regions)):
        for j in regions[i]:
            if j in name.lower():
                return(reg_list[i])
    return -1
#конец скопированной части
def weather_now():
    weather_response = requests.get('http://api.openweathermap.org/data/2.5/weather?q=moscow&appid=cabb0d1a47e3748838dbe5345d78caa9&units=metric&lang=ru')
    info = weather_response.json()
    icon_id = info["weather"][0]["icon"] # иконка погоды
    req_ph = f'http://openweathermap.org/img/wn/' + icon_id + '@2x.png'
    ph = requests.get(req_ph, stream=True).raw
    send_photo(event.user_id, ph, "Погода в Москве")
    w = str(info["weather"][0]["description"].capitalize()) + ", температура: " + str(floor(info["main"]["temp_min"])) + " - " +  str(ceil(info["main"]["temp_max"])) + "°С"
    w += "\nДавление: " + str(round(info["main"]["pressure"]/1.333)) + " мм рт.ст., влажность: " + str(info["main"]["humidity"]) + "%"
    w += "\nВетер: " + wind[float(info['wind']['speed'])] + ", " + str(info['wind']['speed']) + " м/с, " +  wind_direction[float(info['wind']['deg'])]
    send_message(event.user_id, w)
def weather_today():
    w_res = requests.get('https://api.openweathermap.org/data/2.5/forecast?q=moscow&appid=cabb0d1a47e3748838dbe5345d78caa9&units=metric&lang=ru')
    info = w_res.json()
    icons = []
    weather_info = [0]*4
    br_inf, all_inf = '', ''
    for i in range(0, 9, 2):
        if str(info['list'][i]["dt_txt"][:10]) == str(datetime.date.today() + datetime.timedelta(days = 1)) and int(info['list'][i]["dt_txt"][11:13]) > 5:
            break
        weather_info[day_time[(int(info['list'][i]["dt_txt"][11:13]))]] = info['list'][i]
    for j in range(len(weather_info)):
        i = weather_info[j]
        if i != 0:
            icon_id = i["weather"][0]["icon"] # иконка погоды
            req_ph = f'http://openweathermap.org/img/wn/' + icon_id + '@2x.png'
            icons.append(requests.get(req_ph, stream=True))
            br_inf += '/ ' + str(round((float(i['main']['temp_min']) + float(i['main']['temp_min']))/2)) + '°С /'
            all_inf += day_name[j].upper() + '\n'
            all_inf += '//' + str(i["weather"][0]["description"].capitalize()) + ", температура: " + str(floor(i["main"]["temp_min"])) + " - " +  str(ceil(i["main"]["temp_max"])) + "°С"
            all_inf += "\n//Давление: " + str(round(i["main"]["pressure"]/1.333)) + " мм рт.ст., влажность: " + str(i["main"]["humidity"]) + "%"
            all_inf += "\n//Ветер: " + wind[float(i['wind']['speed'])] + ", " + str(i['wind']['speed']) + " м/с, " +  wind_direction[float(i['wind']['deg'])] + '\n'
    img = Image.new('RGBA',(100 * len(icons), 100), color="grey")
    for i in range(len(icons)): # создаем объединенную картинку погоды
        with open(f'file{i+1}.jpeg', 'wb') as f:
            f.write(icons[i].content)
        paste_img = Image.open(f'file{i+1}.jpeg')
        img.paste(paste_img, (i*100, 0), paste_img)
    img.save("today.png")
    send_photo(event.user_id, "today.png", "Погода в Москве сегодня")
    send_message(event.user_id, br_inf)
    send_message(event.user_id, all_inf)
# погода завтра
def weather_tomorrow():
    w_res = requests.get('https://api.openweathermap.org/data/2.5/forecast?q=moscow&appid=cabb0d1a47e3748838dbe5345d78caa9&units=metric&lang=ru')
    info = w_res.json()
    icons = []
    weather_info = [0]*4
    br_inf, all_inf = '', ''
    for i in range(0, 17, 2):
        if str(info['list'][i]["dt_txt"][:10]) == str(datetime.date.today() + datetime.timedelta(days = 2)) and int(info['list'][i]["dt_txt"][11:13]) > 5:
            break
        weather_info[day_time[(int(info['list'][i]["dt_txt"][11:13]))]] = info['list'][i]
    for j in range(len(weather_info)):
        i = weather_info[j]
        if i != 0:
            icon_id = i["weather"][0]["icon"] # иконка погоды
            req_ph = f'http://openweathermap.org/img/wn/' + icon_id + '@2x.png'
            icons.append(requests.get(req_ph, stream=True))
            br_inf += '/ ' + str(round((float(i['main']['temp_min']) + float(i['main']['temp_min']))/2)) + '°С /'
            all_inf += day_name[j].upper() + '\n'
            all_inf += '//' + str(i["weather"][0]["description"].capitalize()) + ", температура: " + str(floor(i["main"]["temp_min"])) + " - " +  str(ceil(i["main"]["temp_max"])) + "°С"
            all_inf += "\n//Давление: " + str(round(i["main"]["pressure"]/1.333)) + " мм рт.ст., влажность: " + str(i["main"]["humidity"]) + "%"
            all_inf += "\n//Ветер: " + wind[float(i['wind']['speed'])] + ", " + str(i['wind']['speed']) + " м/с, " +  wind_direction[float(i['wind']['deg'])] + '\n'
    img = Image.new('RGBA',(100 * len(icons), 100), color="grey")
    for i in range(len(icons)): # создаем объединенную картинку погоды
        with open(f'file{i+1}.jpeg', 'wb') as f:
            f.write(icons[i].content)
        paste_img = Image.open(f'file{i+1}.jpeg')
        img.paste(paste_img, (i*100, 0), paste_img)
    img.save("today.png")
    send_photo(event.user_id, "today.png", "Погода в Москве завтра")
    send_message(event.user_id, br_inf)
    send_message(event.user_id, all_inf)
# прогноз погоды на 5 дней
def weather5():
    w_res = requests.get('https://api.openweathermap.org/data/2.5/forecast?q=moscow&appid=cabb0d1a47e3748838dbe5345d78caa9&units=metric&lang=ru')
    info = w_res.json()
    days = {}
    br_inf, all_inf = '', ''
    for i in range(5):
        days[(datetime.date.today() + datetime.timedelta(days = i)).strftime('%Y-%m-%d')] = [0, 0]
    for i in range(len(info['list'])):
        if info['list'][i]["dt_txt"][:10] == (datetime.date.today() + datetime.timedelta(days = 5)).strftime('%Y-%m-%d'):
            break
        if int(info['list'][i]["dt_txt"][11:13]) >= 6 and int(info['list'][i]["dt_txt"][11:13]) < 17:
            days[info['list'][i]["dt_txt"][:10]][0] = info['list'][i]
        if int(info['list'][i]["dt_txt"][11:13]) >= 21 or int(info['list'][i]["dt_txt"][11:13]) < 4:
            days[info['list'][i]["dt_txt"][:10]][1] = info['list'][i]
    morning, night = '', ''
    icons = []
    print(days)
    for i in days:
        if days[i][0] == 0:
            morning += '/---/'
            if days[i][1] != 0:
                icon_id = days[i][1]["weather"][0]["icon"] # иконка погоды
                req_ph = f'http://openweathermap.org/img/wn/' + icon_id + '@2x.png'
                icons.append(requests.get(req_ph, stream=True))
        else:
            morning += '/ ' + str(ceil(days[i][0]["main"]["temp"])) + '°С /'
            icon_id = days[i][0]["weather"][0]["icon"] # иконка погоды
            req_ph = f'http://openweathermap.org/img/wn/' + icon_id + '@2x.png'
            icons.append(requests.get(req_ph, stream=True))
        if days[i][1] == 0:
            night += '/---/'
        else:
            night += '/ ' + str(ceil(days[i][1]["main"]["temp"])) + '°С /'
    img = Image.new('RGBA',(100 * len(icons), 100), color="grey")
    for i in range(len(icons)): # создаем объединенную картинку погоды
        with open(f'file-5days{i+1}.jpeg', 'wb') as f:
            f.write(icons[i].content)
        paste_img = Image.open(f'file-5days{i+1}.jpeg')
        img.paste(paste_img, (i*100, 0), paste_img)
    img.save("5days.png")
    morning = morning + "ДЕНЬ"
    night = night + "НОЧЬ"
    send_photo(event.user_id, "5days.png", "Погода в Москве с " + (datetime.date.today()).strftime('%d.%m') + ' по ' + (datetime.date.today() + datetime.timedelta(days = 4)).strftime('%d.%m'))
    send_message(event.user_id, morning+'\n'+night)



vk_session = vk_api.VkApi(token='9e9910d7a8894b217dc280d865a71902e5d344ed93fa49a8f079bdec78a1fc1bd675c963dc0455f3a462f') #вставить свой api токен
vk = vk_session.get_api()
if os.path.isfile('data.pickle'):
    with open('data.pickle', 'rb') as f:
        users = pickle.load(f)
else:
    users = {}
longpoll = VkLongPoll(vk_session)
for event in longpoll.listen():
    if event.type == VkEventType.MESSAGE_NEW and event.to_me:
        id = event.user_id  # получаем айди пользователя чтобы отправлять ему сообщения
        message = event.text.lower()  # получаем сообщение со строчными буквами чтобы отвечать на команды
        print('New from {}, text = {}'.format(event.user_id, event.text))
        if not (id in users):  # если нет в базе данных
            users[id] = UserInfo()
            with open('data.pickle', 'wb') as f:
                pickle.dump(users, f)
        if message == "выключить эхо":
            if users[id].answering == "echo":
                users[id].answering = ""
            send_message(id, "Эхо выключено")
        elif users[id].answering == "echo":
            send_message(id, event.text)
        elif message == "какая неделя?":
            send_message(id, f"Сейчас идёт {week_number()} неделя")
            users[id].answering = ""
        elif message == "какая группа?":
            send_message(id, f"Расписание для группы {users[id].other_group.upper()}")
            users[id].answering = ""
        elif users[id].answering != "":
            recieve_answer(users[id].answering)
        elif message == "начать":
            send_message(id,
                         "Здравствуйте! Я – F-бот. Чтобы узнать, что я могу, напишите 'помощь' или выберите пункт из меню.")
            if users[id].name == "":
                send_message(id,
                             "Но сначала позвольте спросить, запоминать ли ваше имя?")
                create_keyboard("name")
                users[id].answering = "start_choosing"
            else:
                create_keyboard("start")
        elif users[id].name == "":
            send_message(id, "Здравствуйте! Вашего имени нет в базе данных. Как соизволите к вам обращаться?")
            users[id].answering = "name"
        elif message == "сменить имя":
            send_message(id, "Как соизволите к вам обращаться?")
            users[id].answering = "name"
        elif message == "забудь меня":
            del users[id]
            send_message(id, "Вы были удалены из базы данных. Чтобы вас снова внесли, отправьте любое сообщение")
        elif message == "расписание":
            if users[id].course == "" or users[id].group == "":
                send_message(id,
                             "Введите свой курс обучения и группу через пробел (курс от 1 до 3 цифрами). Например:\n1 ИКБО-08-21")
                users[id].answering = "group"
            else:
                users[id].answering = "wait_raspis"
                create_keyboard("raspis")
        elif message == "помощь" or "что ты умеешь" in message or message == "команды":
            send_message(id,
                         "На данный момент имеются следующие команды: \n'начать' – отправить приветственное сообщение; "
                         "\n'забудь меня – удалить пользователя из базы данных; \n'сменить имя' – поменять имя, как я буду вас "
                         "называть\n'бот [день недели]' – бот выдаёт расписание на определённый день недели\n'бот [группа]' – выдаёт расписание для определённой группы, например бот икбо-08-20 откроет меню для этой группы \n"
                         "'погода' – прогноз погоды на определённый промежуток времени\n'корона' – статистика коронавируса по России \n'корона [регион]' – статистика для определённого региона ")
        elif "бот" in message:
            users[id].answering = "wait_raspis"
            recieve_answer(users[id].answering)
        elif "найти" in message:
            users[id].answering = "teacher"
            recieve_answer(users[id].answering)
        elif message == "корона":
            #corona_rus()
            send_message(id,"Попытка построить график...")
            try:
                corona_rus()
            except AttributeError:
                send_message(id, "Не удалось получить информацию, попробуйте заново.")
        elif message.split(' ')[0] == "корона":
            corona_reg(' '.join(message.split(' ')[1:]))
        elif message == "погода":
            create_keyboard("weather")
            users[id].answering = "weather"
        else:
            #send_message(id, "Привет, " + users[id].name)
            send_message(id, "Команда не найдена, " + users[id].name)